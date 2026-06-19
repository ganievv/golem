import asyncio
import json
import io
import logging
import os
from datetime import datetime

from aiohttp import web
import httpx
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

BOT_TOKEN  = os.environ["BOT_TOKEN"]
WEBAPP_URL = "https://ganievv.github.io/golem/"
API        = f"https://api.telegram.org/bot{BOT_TOKEN}"
PORT       = int(os.environ.get("PORT", 8080))

logging.basicConfig(format="%(asctime)s - %(levelname)s - %(message)s", level=logging.INFO)
log = logging.getLogger(__name__)


async def send_message(client: httpx.AsyncClient, chat_id: int, text: str, reply_markup=None):
    data: dict = {"chat_id": chat_id, "text": text}
    if reply_markup:
        data["reply_markup"] = reply_markup
    await client.post(f"{API}/sendMessage", json=data)


async def send_document(client: httpx.AsyncClient, chat_id: int,
                        filename: str, content: bytes, caption: str):
    await client.post(
        f"{API}/sendDocument",
        data={"chat_id": chat_id, "caption": caption},
        files={"document": (filename, content,
               "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )


def build_xlsx(orders: list, tips: list) -> tuple[bytes, float, float]:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Заказы"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
    center      = Alignment(horizontal="center", vertical="center")

    ws.append(["Дата", "Стол", "Гость", "Состав заказа", "Сумма (€)", "Чаевые (€)"])
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    for col, w in zip("ABCDEF", [18, 8, 14, 44, 13, 13]):
        ws.column_dimensions[col].width = w

    grand_total = grand_tips = 0.0
    for idx, order in enumerate(orders):
        items_str = ", ".join(f"{it['n']} × {it['q']}" for it in order.get("i", []))
        tip   = float(tips[idx]) if idx < len(tips) else 0.0
        total = float(order.get("s", 0))
        grand_total += total
        grand_tips  += tip
        ws.append([order.get("d", ""), order.get("t", ""), order.get("c", ""),
                   items_str, round(total, 2), round(tip, 2)])

    ws.append([])
    ws.append(["", "", "", "Итого:", round(grand_total, 2), round(grand_tips, 2)])
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True, size=11)

    for row in ws.iter_rows(min_row=2, min_col=5, max_col=6):
        for cell in row:
            cell.number_format = '#,##0.00 €'

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), grand_total, grand_tips


def pluralize(n: int) -> str:
    if n % 10 == 1 and n % 100 != 11:       return f"{n} заказ"
    if n % 10 in (2,3,4) and n % 100 not in (12,13,14): return f"{n} заказа"
    return f"{n} заказов"


async def handle_update(update: dict):
    msg = update.get("message", {})
    if not msg:
        return

    chat_id = msg["chat"]["id"]

    if msg.get("text") == "/start":
        async with httpx.AsyncClient() as client:
            await send_message(client, chat_id, "Нажми кнопку чтобы открыть приложение 👇",
                reply_markup={
                    "inline_keyboard": [[{
                        "text": "🍺 Открыть Bar App",
                        "web_app": {"url": WEBAPP_URL},
                    }]]
                })
        return

    web_app_data = msg.get("web_app_data", {}).get("data")
    if not web_app_data:
        return

    async with httpx.AsyncClient() as client:
        try:
            data   = json.loads(web_app_data)
            orders = data.get("orders", [])
            tips   = data.get("tips", [])
        except json.JSONDecodeError:
            await send_message(client, chat_id, "❌ Ошибка: неверный формат данных.")
            return

        if not orders:
            await send_message(client, chat_id, "Нет заказов для экспорта.")
            return

        content, grand_total, grand_tips = build_xlsx(orders, tips)
        date_str = datetime.now().strftime("%Y-%m-%d")
        filename = f"orders_{date_str}.xlsx"
        caption  = (f"📊 {pluralize(len(orders))}\n"
                    f"💶 Итого: {grand_total:.2f}€\n"
                    f"💰 Чаевые: {grand_tips:.2f}€")

        await send_document(client, chat_id, filename, content, caption)
        log.info("Sent %s to chat %s", filename, chat_id)


async def webhook_handler(request: web.Request) -> web.Response:
    try:
        update = await request.json()
        asyncio.create_task(handle_update(update))
    except Exception as e:
        log.error("Webhook error: %s", e)
    return web.Response(text="ok")


async def register_webhook(app: web.Application):
    hostname = os.environ.get("RENDER_EXTERNAL_HOSTNAME")
    if not hostname:
        log.warning("RENDER_EXTERNAL_HOSTNAME не задан — webhook не зарегистрирован")
        return
    webhook_url = f"https://{hostname}/{BOT_TOKEN}"
    async with httpx.AsyncClient() as client:
        r = await client.post(f"{API}/setWebhook", json={"url": webhook_url})
        log.info("Webhook: %s", r.json())


if __name__ == "__main__":
    aioapp = web.Application()
    aioapp.router.add_post(f"/{BOT_TOKEN}", webhook_handler)
    aioapp.on_startup.append(register_webhook)
    web.run_app(aioapp, port=PORT)
